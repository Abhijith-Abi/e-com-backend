import csv
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


from django.db.models import Count
from django.db.models.functions import TruncMonth, TruncDate
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema, OpenApiResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.analytics.api_v1.dashboard import _parse_date_range
from apps.analytics.api_v1.serializers import ShipmentReportResponseSerializer
from apps.orders.models import Order
from core.choices import CurrencyChoices, OrderStatusChoices


_PARAMS = [
    OpenApiParameter("date_from", description="Start date (YYYY-MM-DD)", required=False),
    OpenApiParameter("date_to", description="End date (YYYY-MM-DD)", required=False),
    OpenApiParameter("from_date", description="Alternative start date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter("to_date", description="Alternative end date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter("currency", description="Filter by currency: INR, USD, GBP", required=False, enum=["INR", "USD", "GBP"]),
    OpenApiParameter("dummy", description="Return dummy data (dummy=true)", required=False),
]


def _build_shipment_report(date_from, date_to, currency=None, warehouse=None):
    wh_filter = {"warehouse": warehouse} if warehouse else {}
    currency_filter = {"currency": currency} if currency else {}

    base_qs = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        is_deleted=False,
        **wh_filter,
        **currency_filter,
    )

    total_shipped = base_qs.filter(
        order_status__in=[OrderStatusChoices.SHIPPED, OrderStatusChoices.DELIVERED]
    ).count()
    total_delivered = base_qs.filter(order_status=OrderStatusChoices.DELIVERED).count()
    in_transit = base_qs.filter(order_status=OrderStatusChoices.SHIPPED).count()

    # Daily trend: shipped vs delivered per day
    daily_qs = (
        base_qs.filter(
            order_status__in=[OrderStatusChoices.SHIPPED, OrderStatusChoices.DELIVERED]
        )
        .annotate(date=TruncDate("created_at"))
        .values("date", "order_status")
        .annotate(count=Count("id"))
        .order_by("date", "order_status")
    )

    daily_map = {}
    for row in daily_qs:
        date_str = row["date"].strftime("%Y-%m-%d") if row["date"] else ""
        if not date_str:
            continue
        if date_str not in daily_map:
            daily_map[date_str] = {"shipped": 0, "delivered": 0}
        if row["order_status"] == OrderStatusChoices.SHIPPED:
            daily_map[date_str]["shipped"] += row["count"]
        elif row["order_status"] == OrderStatusChoices.DELIVERED:
            daily_map[date_str]["delivered"] += row["count"]

    monthly_chart = [
        {"month": date_key, "shipped": v["shipped"], "delivered": v["delivered"]}
        for date_key, v in sorted(daily_map.items())
    ]

    return {
        "date_from": date_from.date(),
        "date_to": date_to.date(),
        "currency_filter": currency,
        "warehouse_id": str(warehouse.id) if warehouse else None,
        "total_shipped": total_shipped,
        "total_delivered": total_delivered,
        "in_transit": in_transit,
        "monthly_shipment_chart": monthly_chart,
    }


def _build_dummy_shipment_report(currency=None):
    dates = ["2026-05-20", "2026-05-21", "2026-05-22", "2026-05-23", "2026-05-24", "2026-05-25"]
    shipped = [32, 31, 37, 38, 42, 42]
    delivered = [30, 29, 35, 36, 40, 41]
    return {
        "date_from": "2026-05-20",
        "date_to": "2026-05-25",
        "currency_filter": currency,
        "warehouse_id": None,
        "total_shipped": 222,
        "total_delivered": 211,
        "in_transit": 8,
        "monthly_shipment_chart": [
            {"month": d, "shipped": s, "delivered": d_val}
            for d, s, d_val in zip(dates, shipped, delivered)
        ],
    }


@extend_schema(tags=["Analytics"], summary="Shipment report — shipped, delivered, in-transit counts + monthly trend", parameters=_PARAMS, responses={200: ShipmentReportResponseSerializer})
class ShipmentReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            return Response(_build_dummy_shipment_report(currency))

        date_from, date_to = _parse_date_range(request)
        return Response(_build_shipment_report(date_from, date_to, currency=currency))


@extend_schema(tags=["Warehouse Analytics"], summary="Warehouse shipment report — shipped, delivered, in-transit counts + monthly trend", parameters=_PARAMS, responses={200: ShipmentReportResponseSerializer})
class WarehouseShipmentReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            data = _build_dummy_shipment_report(currency)
            wh = getattr(request, "warehouse", None)
            data["warehouse_id"] = str(wh.id) if wh else None
            return Response(data)

        date_from, date_to = _parse_date_range(request)
        return Response(_build_shipment_report(date_from, date_to, currency=currency, warehouse=request.warehouse))



class _ShipmentExportMixin:
    def _export_csv(self, data):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Shipment Report"])
        writer.writerow(["Date From", data["date_from"], "Date To", data["date_to"]])
        writer.writerow([])
        writer.writerow(["Total Shipped", data["total_shipped"]])
        writer.writerow(["Total Delivered", data["total_delivered"]])
        writer.writerow(["In Transit", data["in_transit"]])
        writer.writerow([])
        writer.writerow(["Daily Shipment Trend"])
        writer.writerow(["Date", "Shipped", "Delivered"])
        for row in data["monthly_shipment_chart"]:
            writer.writerow([row["month"], row["shipped"], row["delivered"]])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="shipment_report.csv"'
        return response

    def _export_excel(self, data):
        if not EXCEL_AVAILABLE:
            return Response(
                {"detail": "Excel export not available. Install openpyxl: pip install openpyxl"},
                status=400
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shipment Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="4472C4")
        section_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        
        row = 1

        # Title
        ws.merge_cells(f"A{row}:D{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "SHIPMENT REPORT"
        title_cell.font = title_font
        title_cell.alignment = center_aligned
        row += 2

        # Date range
        ws[f"A{row}"] = "Date From:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = str(data["date_from"])
        ws[f"B{row}"].alignment = left_aligned
        
        ws[f"C{row}"] = "Date To:"
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"D{row}"] = str(data["date_to"])
        ws[f"D{row}"].alignment = left_aligned
        row += 2

        # --- Report Summary Table ---
        ws[f"A{row}"] = "Report Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        for col_idx, header in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        row += 1

        metrics = [
            ("Total Shipped", data["total_shipped"]),
            ("Total Delivered", data["total_delivered"]),
            ("In Transit", data["in_transit"]),
        ]

        for i, (metric, value) in enumerate(metrics):
            cell_metric = ws.cell(row=row, column=1)
            cell_metric.value = metric
            cell_metric.border = border
            cell_metric.alignment = left_aligned
            
            cell_value = ws.cell(row=row, column=2)
            cell_value.value = value
            cell_value.border = border
            cell_value.alignment = right_aligned
            
            if i % 2 == 1:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_metric.fill = fill
                cell_value.fill = fill
                
            row += 1
            
        row += 2

        # --- Daily Chart Table ---
        ws[f"A{row}"] = "Daily Shipment Trend"
        ws[f"A{row}"].font = section_font
        row += 1
 
        headers = ["Date", "Shipped", "Delivered"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        
        row += 1

        for i, chart_row in enumerate(data["monthly_shipment_chart"]):
            for col_idx, key in enumerate(["month", "shipped", "delivered"], 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = chart_row.get(key, "")
                cell.border = border
                
                if col_idx == 1:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = right_aligned
                    
                if i % 2 == 1:
                    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.fill = fill
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="shipment_report.xlsx"'
        return response


@extend_schema(tags=["Analytics"], summary="Export shipment report", parameters=_PARAMS)
class ShipmentReportExportView(APIView, _ShipmentExportMixin):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_shipment_report(date_from, date_to, currency=currency)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)


@extend_schema(tags=["Warehouse Analytics"], summary="Export warehouse shipment report", parameters=_PARAMS)
class WarehouseShipmentReportExportView(APIView, _ShipmentExportMixin):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_shipment_report(date_from, date_to, currency=currency, warehouse=request.warehouse)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)
