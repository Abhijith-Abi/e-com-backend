import csv
import io

from django.db.models import DecimalField, Sum
from django.db.models.functions import Coalesce, TruncMonth, TruncDate
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema, OpenApiResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.analytics.api_v1.dashboard import _parse_date_range
from apps.analytics.api_v1.serializers import SalesReportResponseSerializer
from apps.orders.models import Order
from core.choices import CurrencyChoices, OrderStatusChoices

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


_SALES_REPORT_PARAMS = [
    OpenApiParameter("date_from", description="Start date (YYYY-MM-DD)", required=False),
    OpenApiParameter("date_to", description="End date (YYYY-MM-DD)", required=False),
    OpenApiParameter("from_date", description="Alternative start date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter("to_date", description="Alternative end date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter(
        "currency",
        description="Filter by currency: INR, USD, GBP. Omit for all.",
        required=False,
        enum=["INR", "USD", "GBP"],
    ),
    OpenApiParameter("dummy", description="Return dummy data (dummy=true)", required=False),
]


def _build_sales_report(date_from, date_to, currency=None, warehouse=None):
    wh_filter = {"warehouse": warehouse} if warehouse else {}
    currency_filter = {"currency": currency} if currency else {}

    # Include all orders (not just paid) since we use COD
    orders_qs = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        is_deleted=False,
        **wh_filter,
        **currency_filter,
    )

    # Total revenue per currency
    revenue_by_currency = {}
    for cur in CurrencyChoices.values:
        if currency and cur != currency:
            revenue_by_currency[cur] = None  # not requested
            continue
        total = orders_qs.filter(currency=cur).aggregate(
            total=Coalesce(Sum("total_amount"), 0, output_field=DecimalField())
        )["total"]
        revenue_by_currency[cur] = float(total)

    # Daily revenue chart (grouped by date, per currency or combined)
    daily_qs = (
        orders_qs
        .annotate(date=TruncDate("created_at"))
        .values("date", "currency")
        .annotate(revenue=Coalesce(Sum("total_amount"), 0, output_field=DecimalField()))
        .order_by("date", "currency")
    )

    # Build a dict: {date_str: {currency: revenue}}
    daily_map = {}
    for row in daily_qs:
        date_str = row["date"].strftime("%Y-%m-%d") if row["date"] else ""
        if not date_str:
            continue
        if date_str not in daily_map:
            daily_map[date_str] = {}
        daily_map[date_str][row["currency"]] = float(row["revenue"])

    monthly_chart = [
        {"month": date_key, **revenues}
        for date_key, revenues in daily_map.items()
    ]

    # Summary counts
    total_orders = orders_qs.count()
    cancelled_orders = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        order_status=OrderStatusChoices.CANCELLED,
        is_deleted=False,
        **wh_filter,
        **currency_filter,
    ).count()

    all_orders_count = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        is_deleted=False,
        **wh_filter,
        **currency_filter,
    ).count()

    cancellation_rate = (
        round((cancelled_orders / all_orders_count) * 100, 1) if all_orders_count else 0
    )

    return {
        "date_from": date_from.date(),
        "date_to": date_to.date(),
        "currency_filter": currency,
        "warehouse_id": str(warehouse.id) if warehouse else None,
        "total_revenue": {
            cur: val for cur, val in revenue_by_currency.items() if val is not None
        },
        "total_paid_orders": total_orders,
        "cancelled_orders": cancelled_orders,
        "cancellation_rate_pct": cancellation_rate,
        "monthly_revenue_chart": monthly_chart,
    }


def _build_dummy_sales_report(currency=None):
    dates = ["2026-05-20", "2026-05-21", "2026-05-22", "2026-05-23", "2026-05-24", "2026-05-25"]
    inr_vals = [42000, 38000, 49000, 45000, 57000, 61000]
    usd_vals = [510, 460, 590, 540, 690, 740]
    gbp_vals = [420, 380, 490, 450, 570, 610]

    if currency == "INR":
        chart = [{"month": d, "INR": v} for d, v in zip(dates, inr_vals)]
        total_revenue = {"INR": sum(inr_vals)}
    elif currency == "USD":
        chart = [{"month": d, "USD": v} for d, v in zip(dates, usd_vals)]
        total_revenue = {"USD": sum(usd_vals)}
    elif currency == "GBP":
        chart = [{"month": d, "GBP": v} for d, v in zip(dates, gbp_vals)]
        total_revenue = {"GBP": sum(gbp_vals)}
    else:
        chart = [
            {"month": d, "INR": i, "USD": u, "GBP": g}
            for d, i, u, g in zip(dates, inr_vals, usd_vals, gbp_vals)
        ]
        total_revenue = {"INR": sum(inr_vals), "USD": sum(usd_vals), "GBP": sum(gbp_vals)}

    return {
        "date_from": "2026-05-20",
        "date_to": "2026-05-25",
        "currency_filter": currency,
        "warehouse_id": None,
        "total_revenue": total_revenue,
        "total_paid_orders": 84,
        "cancelled_orders": 3,
        "cancellation_rate_pct": 3.6,
        "monthly_revenue_chart": chart,
    }


@extend_schema(
    tags=["Analytics"],
    summary="Sales report — total revenue by currency + monthly chart",
    parameters=_SALES_REPORT_PARAMS,
    responses={200: SalesReportResponseSerializer},
)
class SalesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            return Response(_build_dummy_sales_report(currency))

        date_from, date_to = _parse_date_range(request)
        return Response(_build_sales_report(date_from, date_to, currency=currency))


@extend_schema(
    tags=["Warehouse Analytics"],
    summary="Warehouse sales report — total revenue by currency + monthly chart",
    parameters=_SALES_REPORT_PARAMS,
    responses={200: SalesReportResponseSerializer},
)
class WarehouseSalesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        print('HIT GET')
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            wh = getattr(request, "warehouse", None)
            data = _build_dummy_sales_report(currency)
            data["warehouse_id"] = str(wh.id) if wh else None
            return Response(data)

        date_from, date_to = _parse_date_range(request)
        return Response(_build_sales_report(date_from, date_to, currency=currency, warehouse=request.warehouse))


@extend_schema(
    tags=["Analytics"],
    summary="Export sales report as CSV or Excel",
    parameters=_SALES_REPORT_PARAMS + [
        OpenApiParameter("format", description="Export format: csv or excel (default: csv)", required=False, enum=["csv", "excel"]),
    ],
    responses={200: OpenApiResponse(description="File attachment (CSV or Excel)")},
)
class SalesReportExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_sales_report(date_from, date_to, currency=currency)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)

    def _export_csv(self, data):
        """Export as CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Summary section
        writer.writerow(["Sales Report"])
        writer.writerow(["Date From", data["date_from"], "Date To", data["date_to"]])
        writer.writerow([])
        headers = [f"Total Revenue ({cur})" for cur in data["total_revenue"].keys()] + ["Total Paid Orders", "Cancelled Orders", "Cancellation Rate (%)"]
        writer.writerow(headers)
        values = list(data["total_revenue"].values()) + [data["total_paid_orders"], data["cancelled_orders"], data["cancellation_rate_pct"]]
        writer.writerow(values)
        writer.writerow([])

        # Daily chart section
        writer.writerow(["Daily Revenue"])
        if data["monthly_revenue_chart"]:
            headers = list(data["monthly_revenue_chart"][0].keys())
            writer.writerow(headers)
            for row in data["monthly_revenue_chart"]:
                writer.writerow([row.get(h, "") for h in headers])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="sales_report.csv"'
        return response

    def _export_excel(self, data):
        """Export as Excel"""
        if not EXCEL_AVAILABLE:
            return Response(
                {"detail": "Excel export not available. Install openpyxl: pip install openpyxl"},
                status=400
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="4472C4")
        section_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        
        row = 1

        # Title
        ws.merge_cells(f"A{row}:D{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "SALES REPORT"
        title_cell.font = title_font
        title_cell.alignment = center_aligned
        row += 2

        # Date range
        ws[f"A{row}"] = "Date From:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = str(data["date_from"])
        ws[f"B{row}"].alignment = left_aligned
        
        ws[f"C{row}"] = "Date To:"
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"D{row}"] = str(data["date_to"])
        ws[f"D{row}"].alignment = left_aligned
        row += 2

        # --- Report Summary Table ---
        ws[f"A{row}"] = "Report Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        # Table headers
        for col_idx, header in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        row += 1

        # Table data
        metrics = []
        for cur, val in data["total_revenue"].items():
            metrics.append((f"Total Revenue ({cur})", val))
        metrics.extend([
            ("Total Paid Orders", data["total_paid_orders"]),
            ("Cancelled Orders", data["cancelled_orders"]),
            ("Cancellation Rate (%)", data["cancellation_rate_pct"]),
        ])

        for i, (metric, value) in enumerate(metrics):
            cell_metric = ws.cell(row=row, column=1)
            cell_metric.value = metric
            cell_metric.border = border
            cell_metric.alignment = left_aligned
            
            cell_value = ws.cell(row=row, column=2)
            cell_value.value = value
            cell_value.border = border
            cell_value.alignment = right_aligned
            
            # Alternate row shading
            if i % 2 == 1:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_metric.fill = fill
                cell_value.fill = fill
                
            row += 1
            
        row += 2

        # --- Daily Revenue Table ---
        ws[f"A{row}"] = "Daily Revenue"
        ws[f"A{row}"].font = section_font
        row += 1

        if data["monthly_revenue_chart"]:
            headers = list(data["monthly_revenue_chart"][0].keys())
            
            # Header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                # Capitalize header (e.g. 'month' -> 'Month', 'INR' -> 'INR')
                cell.value = header.capitalize() if header.islower() else header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_aligned
            
            row += 1

            # Data rows
            for i, chart_row in enumerate(data["monthly_revenue_chart"]):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = chart_row.get(header, "")
                    cell.border = border
                    
                    if col_idx == 1:
                        cell.alignment = left_aligned
                    else:
                        cell.alignment = right_aligned
                        
                    # Alternate row shading
                    if i % 2 == 1:
                        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.fill = fill
                        
                row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'
        return response



@extend_schema(
    tags=["Warehouse Analytics"],
    summary="Warehouse sales report export as CSV or Excel",
    parameters=_SALES_REPORT_PARAMS + [
        OpenApiParameter("format", description="Export format: csv or excel (default: csv)", required=False, enum=["csv", "excel"]),
    ],
    responses={200: OpenApiResponse(description="File attachment (CSV or Excel)")},
)
class WarehouseSalesReportExportView(APIView):
    def dispatch(self, request, *args, **kwargs):
        print('HIT DISPATCH')
        return super().dispatch(request, *args, **kwargs)
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        print('HIT GET')
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_sales_report(date_from, date_to, currency=currency, warehouse=request.warehouse)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)

    def _export_csv(self, data):
        """Export as CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Summary section
        writer.writerow(["Sales Report"])
        writer.writerow(["Date From", data["date_from"], "Date To", data["date_to"]])
        writer.writerow([])
        headers = [f"Total Revenue ({cur})" for cur in data["total_revenue"].keys()] + ["Total Paid Orders", "Cancelled Orders", "Cancellation Rate (%)"]
        writer.writerow(headers)
        values = list(data["total_revenue"].values()) + [data["total_paid_orders"], data["cancelled_orders"], data["cancellation_rate_pct"]]
        writer.writerow(values)
        writer.writerow([])

        # Daily chart section
        writer.writerow(["Daily Revenue"])
        if data["monthly_revenue_chart"]:
            headers = list(data["monthly_revenue_chart"][0].keys())
            writer.writerow(headers)
            for row in data["monthly_revenue_chart"]:
                writer.writerow([row.get(h, "") for h in headers])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="sales_report.csv"'
        return response

    def _export_excel(self, data):
        """Export as Excel"""
        if not EXCEL_AVAILABLE:
            return Response(
                {"detail": "Excel export not available. Install openpyxl: pip install openpyxl"},
                status=400
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="4472C4")
        section_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        
        row = 1

        # Title
        ws.merge_cells(f"A{row}:D{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "SALES REPORT"
        title_cell.font = title_font
        title_cell.alignment = center_aligned
        row += 2

        # Date range
        ws[f"A{row}"] = "Date From:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = str(data["date_from"])
        ws[f"B{row}"].alignment = left_aligned
        
        ws[f"C{row}"] = "Date To:"
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"D{row}"] = str(data["date_to"])
        ws[f"D{row}"].alignment = left_aligned
        row += 2

        # --- Report Summary Table ---
        ws[f"A{row}"] = "Report Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        # Table headers
        for col_idx, header in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        row += 1

        # Table data
        metrics = []
        for cur, val in data["total_revenue"].items():
            metrics.append((f"Total Revenue ({cur})", val))
        metrics.extend([
            ("Total Paid Orders", data["total_paid_orders"]),
            ("Cancelled Orders", data["cancelled_orders"]),
            ("Cancellation Rate (%)", data["cancellation_rate_pct"]),
        ])

        for i, (metric, value) in enumerate(metrics):
            cell_metric = ws.cell(row=row, column=1)
            cell_metric.value = metric
            cell_metric.border = border
            cell_metric.alignment = left_aligned
            
            cell_value = ws.cell(row=row, column=2)
            cell_value.value = value
            cell_value.border = border
            cell_value.alignment = right_aligned
            
            # Alternate row shading
            if i % 2 == 1:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_metric.fill = fill
                cell_value.fill = fill
                
            row += 1
            
        row += 2

        # --- Daily Revenue Table ---
        ws[f"A{row}"] = "Daily Revenue"
        ws[f"A{row}"].font = section_font
        row += 1

        if data["monthly_revenue_chart"]:
            headers = list(data["monthly_revenue_chart"][0].keys())
            
            # Header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                # Capitalize header (e.g. 'month' -> 'Month', 'INR' -> 'INR')
                cell.value = header.capitalize() if header.islower() else header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_aligned
            
            row += 1

            # Data rows
            for i, chart_row in enumerate(data["monthly_revenue_chart"]):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = chart_row.get(header, "")
                    cell.border = border
                    
                    if col_idx == 1:
                        cell.alignment = left_aligned
                    else:
                        cell.alignment = right_aligned
                        
                    # Alternate row shading
                    if i % 2 == 1:
                        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.fill = fill
                        
                row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'
        return response
