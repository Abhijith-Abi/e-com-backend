import csv
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema, OpenApiResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.analytics.api_v1.dashboard import _parse_date_range
from apps.analytics.api_v1.serializers import OrdersReportResponseSerializer
from apps.orders.models import Order
from core.choices import CurrencyChoices, OrderStatusChoices


_ORDERS_REPORT_PARAMS = [
    OpenApiParameter("date_from", description="Start date (YYYY-MM-DD)", required=False),
    OpenApiParameter("date_to", description="End date (YYYY-MM-DD)", required=False),
    OpenApiParameter("from_date", description="Alternative start date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter("to_date", description="Alternative end date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter(
        "currency",
        description="Filter by currency: INR, USD, GBP. Omit for all.",
        required=False,
        enum=["INR", "USD", "GBP"],
    ),
    OpenApiParameter("dummy", description="Return dummy data (dummy=true)", required=False),
]

_ALL_STATUSES = [
    OrderStatusChoices.DELIVERED,
    OrderStatusChoices.SHIPPED,
    OrderStatusChoices.PROCESSING,
    OrderStatusChoices.PENDING,
    OrderStatusChoices.CANCELLED,
    OrderStatusChoices.CONFIRMED,
]


def _build_orders_report(date_from, date_to, currency=None, warehouse=None):
    wh_filter = {"warehouse": warehouse} if warehouse else {}
    currency_filter = {"currency": currency} if currency else {}

    qs = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        is_deleted=False,
        **wh_filter,
        **currency_filter,
    )

    counts_qs = (
        qs.values("order_status")
        .annotate(count=Count("id"))
    )
    counts_map = {row["order_status"]: row["count"] for row in counts_qs}

    by_status = [
        {"status": status, "count": counts_map.get(status, 0)}
        for status in _ALL_STATUSES
    ]

    total = sum(row["count"] for row in by_status)

    daily_qs = (
        qs.annotate(date=TruncDate("created_at"))
        .values("date")
        .annotate(count=Count("id"))
        .order_by("date")
    )
    daily_orders_chart = [
        {"date": row["date"].strftime("%Y-%m-%d") if row["date"] else "", "count": row["count"]}
        for row in daily_qs
        if row["date"]
    ]

    return {
        "date_from": date_from.date(),
        "date_to": date_to.date(),
        "currency_filter": currency,
        "warehouse_id": str(warehouse.id) if warehouse else None,
        "total_orders": total,
        "by_status": by_status,
        "daily_orders_chart": daily_orders_chart,
    }


def _build_dummy_orders_report(currency=None):
    by_status = [
        {"status": "delivered", "count": 1245},
        {"status": "shipped", "count": 342},
        {"status": "processing", "count": 186},
        {"status": "pending", "count": 43},
        {"status": "cancelled", "count": 89},
        {"status": "confirmed", "count": 0},
    ]
    dates = ["2026-05-20", "2026-05-21", "2026-05-22", "2026-05-23", "2026-05-24", "2026-05-25"]
    daily_counts = [10, 15, 8, 22, 14, 15]
    daily_orders_chart = [{"date": d, "count": c} for d, c in zip(dates, daily_counts)]
    return {
        "date_from": "2026-05-20",
        "date_to": "2026-05-25",
        "currency_filter": currency,
        "warehouse_id": None,
        "total_orders": sum(r["count"] for r in by_status),
        "by_status": by_status,
        "daily_orders_chart": daily_orders_chart,
    }


@extend_schema(
    tags=["Analytics"],
    summary="Orders report — order counts by status",
    parameters=_ORDERS_REPORT_PARAMS,
    responses={200: OrdersReportResponseSerializer},
)
class OrdersReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            return Response(_build_dummy_orders_report(currency))

        date_from, date_to = _parse_date_range(request)
        return Response(_build_orders_report(date_from, date_to, currency=currency))


@extend_schema(
    tags=["Warehouse Analytics"],
    summary="Warehouse orders report — order counts by status",
    parameters=_ORDERS_REPORT_PARAMS,
    responses={200: OrdersReportResponseSerializer},
)
class WarehouseOrdersReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            data = _build_dummy_orders_report(currency)
            wh = getattr(request, "warehouse", None)
            data["warehouse_id"] = str(wh.id) if wh else None
            return Response(data)

        date_from, date_to = _parse_date_range(request)
        return Response(_build_orders_report(date_from, date_to, currency=currency, warehouse=request.warehouse))



class _OrdersExportMixin:
    def _export_csv(self, data):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Orders Report"])
        writer.writerow(["Date From", data["date_from"], "Date To", data["date_to"]])
        writer.writerow(["Total Orders", data["total_orders"]])
        writer.writerow([])
        writer.writerow(["Status", "Count"])
        for row in data["by_status"]:
            writer.writerow([row["status"].capitalize(), row["count"]])
        
        writer.writerow([])
        writer.writerow(["Daily Order Trend"])
        writer.writerow(["Date", "Count"])
        for row in data.get("daily_orders_chart", []):
            writer.writerow([row["date"], row["count"]])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="orders_report.csv"'
        return response

    def _export_excel(self, data):
        if not EXCEL_AVAILABLE:
            return Response(
                {"detail": "Excel export not available. Install openpyxl: pip install openpyxl"},
                status=400
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="4472C4")
        section_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        
        row = 1

        # Title
        ws.merge_cells(f"A{row}:D{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "ORDERS REPORT"
        title_cell.font = title_font
        title_cell.alignment = center_aligned
        row += 2

        # Date range
        ws[f"A{row}"] = "Date From:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = str(data["date_from"])
        ws[f"B{row}"].alignment = left_aligned
        
        ws[f"C{row}"] = "Date To:"
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"D{row}"] = str(data["date_to"])
        ws[f"D{row}"].alignment = left_aligned
        row += 2

        # --- Report Summary Table ---
        ws[f"A{row}"] = "Total Orders Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        for col_idx, header in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        row += 1

        cell_metric = ws.cell(row=row, column=1)
        cell_metric.value = "Total Orders"
        cell_metric.border = border
        cell_metric.alignment = left_aligned
        
        cell_value = ws.cell(row=row, column=2)
        cell_value.value = data["total_orders"]
        cell_value.border = border
        cell_value.alignment = right_aligned
        
        row += 2

        # --- Orders By Status Table ---
        ws[f"A{row}"] = "Orders by Status"
        ws[f"A{row}"].font = section_font
        row += 1

        for col_idx, header in enumerate(["Status", "Count"], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        
        row += 1

        for i, status_row in enumerate(data["by_status"]):
            cell_status = ws.cell(row=row, column=1)
            cell_status.value = status_row["status"].capitalize()
            cell_status.border = border
            cell_status.alignment = left_aligned
            
            cell_count = ws.cell(row=row, column=2)
            cell_count.value = status_row["count"]
            cell_count.border = border
            cell_count.alignment = right_aligned
            
            if i % 2 == 1:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_status.fill = fill
                cell_count.fill = fill
                
            row += 1

        row += 2

        # --- Daily Orders Chart Table ---
        ws[f"A{row}"] = "Daily Orders Trend"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Date", "Count"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        row += 1

        for i, chart_row in enumerate(data.get("daily_orders_chart", [])):
            cell_date = ws.cell(row=row, column=1)
            cell_date.value = chart_row.get("date", "")
            cell_date.border = border
            cell_date.alignment = left_aligned
            
            cell_count = ws.cell(row=row, column=2)
            cell_count.value = chart_row.get("count", 0)
            cell_count.border = border
            cell_count.alignment = right_aligned
            
            if i % 2 == 1:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_date.fill = fill
                cell_count.fill = fill
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders_report.xlsx"'
        return response


@extend_schema(tags=["Analytics"], summary="Export orders report", parameters=_ORDERS_REPORT_PARAMS)
class OrdersReportExportView(APIView, _OrdersExportMixin):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_orders_report(date_from, date_to, currency=currency)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)


@extend_schema(tags=["Warehouse Analytics"], summary="Export warehouse orders report", parameters=_ORDERS_REPORT_PARAMS)
class WarehouseOrdersReportExportView(APIView, _OrdersExportMixin):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_orders_report(date_from, date_to, currency=currency, warehouse=request.warehouse)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)
