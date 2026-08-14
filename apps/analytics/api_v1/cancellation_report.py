import csv
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


from django.db.models import Count
from django.db.models.functions import TruncMonth, TruncDate
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiParameter, extend_schema, OpenApiResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.analytics.api_v1.dashboard import _parse_date_range
from apps.analytics.api_v1.serializers import CancellationReportResponseSerializer
from apps.orders.models import Order
from core.choices import CurrencyChoices, OrderStatusChoices, PaymentStatusChoices


_PARAMS = [
    OpenApiParameter("date_from", description="Start date (YYYY-MM-DD)", required=False),
    OpenApiParameter("date_to", description="End date (YYYY-MM-DD)", required=False),
    OpenApiParameter("from_date", description="Alternative start date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter("to_date", description="Alternative end date parameter (YYYY-MM-DD)", required=False),
    OpenApiParameter("currency", description="Filter by currency: INR, USD, GBP", required=False, enum=["INR", "USD", "GBP"]),
    OpenApiParameter("dummy", description="Return dummy data (dummy=true)", required=False),
]


def _build_cancellation_report(date_from, date_to, currency=None, warehouse=None):
    wh_filter = {"warehouse": warehouse} if warehouse else {}
    currency_filter = {"currency": currency} if currency else {}

    base_qs = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        is_deleted=False,
        **wh_filter,
        **currency_filter,
    )

    total_cancelled = base_qs.filter(order_status=OrderStatusChoices.CANCELLED).count()
    total_refunded = base_qs.filter(payment_status=PaymentStatusChoices.REFUNDED).count()

    # Daily trend: cancelled vs refunded per day
    cancelled_daily = (
        base_qs.filter(order_status=OrderStatusChoices.CANCELLED)
        .annotate(date=TruncDate("created_at"))
        .values("date")
        .annotate(cancelled=Count("id"))
        .order_by("date")
    )

    refunded_daily = (
        base_qs.filter(payment_status=PaymentStatusChoices.REFUNDED)
        .annotate(date=TruncDate("created_at"))
        .values("date")
        .annotate(refunded=Count("id"))
        .order_by("date")
    )

    daily_map = {}
    for row in cancelled_daily:
        date_str = row["date"].strftime("%Y-%m-%d") if row["date"] else ""
        if date_str:
            daily_map.setdefault(date_str, {"cancelled": 0, "refunded": 0})
            daily_map[date_str]["cancelled"] = row["cancelled"]

    for row in refunded_daily:
        date_str = row["date"].strftime("%Y-%m-%d") if row["date"] else ""
        if date_str:
            daily_map.setdefault(date_str, {"cancelled": 0, "refunded": 0})
            daily_map[date_str]["refunded"] = row["refunded"]

    monthly_chart = [
        {"month": date_key, "cancelled": v["cancelled"], "refunded": v["refunded"]}
        for date_key, v in sorted(daily_map.items())
    ]

    return {
        "date_from": date_from.date(),
        "date_to": date_to.date(),
        "currency_filter": currency,
        "warehouse_id": str(warehouse.id) if warehouse else None,
        "total_cancelled": total_cancelled,
        "total_refunded": total_refunded,
        "monthly_cancellation_chart": monthly_chart,
    }


def _build_dummy_cancellation_report(currency=None):
    dates = ["2026-05-20", "2026-05-21", "2026-05-22", "2026-05-23", "2026-05-24", "2026-05-25"]
    cancelled = [1, 2, 0, 3, 1, 1]
    refunded = [1, 1, 0, 2, 1, 0]
    return {
        "date_from": "2026-05-20",
        "date_to": "2026-05-25",
        "currency_filter": currency,
        "warehouse_id": None,
        "total_cancelled": 8,
        "total_refunded": 5,
        "monthly_cancellation_chart": [
            {"month": d, "cancelled": c, "refunded": r}
            for d, c, r in zip(dates, cancelled, refunded)
        ],
    }


@extend_schema(tags=["Analytics"], summary="Cancellation report — cancelled & refunded counts + monthly trend", parameters=_PARAMS, responses={200: CancellationReportResponseSerializer})
class CancellationReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            return Response(_build_dummy_cancellation_report(currency))

        date_from, date_to = _parse_date_range(request)
        return Response(_build_cancellation_report(date_from, date_to, currency=currency))


@extend_schema(tags=["Warehouse Analytics"], summary="Warehouse cancellation report — cancelled & refunded counts + monthly trend", parameters=_PARAMS, responses={200: CancellationReportResponseSerializer})
class WarehouseCancellationReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": f"Invalid currency. Choose from {CurrencyChoices.values}."}, status=400)

        if request.query_params.get("dummy", "").lower() == "true":
            data = _build_dummy_cancellation_report(currency)
            wh = getattr(request, "warehouse", None)
            data["warehouse_id"] = str(wh.id) if wh else None
            return Response(data)

        date_from, date_to = _parse_date_range(request)
        return Response(_build_cancellation_report(date_from, date_to, currency=currency, warehouse=request.warehouse))



class _CancellationExportMixin:
    def _export_csv(self, data):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Cancellation Report"])
        writer.writerow(["Date From", data["date_from"], "Date To", data["date_to"]])
        writer.writerow([])
        writer.writerow(["Total Cancelled", data["total_cancelled"]])
        writer.writerow(["Total Refunded", data["total_refunded"]])
        writer.writerow([])
        writer.writerow(["Daily Cancellation Trend"])
        writer.writerow(["Date", "Cancelled", "Refunded"])
        for row in data["monthly_cancellation_chart"]:
            writer.writerow([row["month"], row["cancelled"], row["refunded"]])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="cancellation_report.csv"'
        return response

    def _export_excel(self, data):
        if not EXCEL_AVAILABLE:
            return Response(
                {"detail": "Excel export not available. Install openpyxl: pip install openpyxl"},
                status=400
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cancellation Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16, color="4472C4")
        section_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        
        row = 1

        # Title
        ws.merge_cells(f"A{row}:D{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "CANCELLATION REPORT"
        title_cell.font = title_font
        title_cell.alignment = center_aligned
        row += 2

        # Date range
        ws[f"A{row}"] = "Date From:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = str(data["date_from"])
        ws[f"B{row}"].alignment = left_aligned
        
        ws[f"C{row}"] = "Date To:"
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"D{row}"] = str(data["date_to"])
        ws[f"D{row}"].alignment = left_aligned
        row += 2

        # --- Report Summary Table ---
        ws[f"A{row}"] = "Report Summary"
        ws[f"A{row}"].font = section_font
        row += 1

        for col_idx, header in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        row += 1

        metrics = [
            ("Total Cancelled", data["total_cancelled"]),
            ("Total Refunded", data["total_refunded"]),
        ]

        for i, (metric, value) in enumerate(metrics):
            cell_metric = ws.cell(row=row, column=1)
            cell_metric.value = metric
            cell_metric.border = border
            cell_metric.alignment = left_aligned
            
            cell_value = ws.cell(row=row, column=2)
            cell_value.value = value
            cell_value.border = border
            cell_value.alignment = right_aligned
            
            if i % 2 == 1:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_metric.fill = fill
                cell_value.fill = fill
                
            row += 1
            
        row += 2

        # --- Daily Chart Table ---
        ws[f"A{row}"] = "Daily Cancellation Trend"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Date", "Cancelled", "Refunded"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_aligned
        
        row += 1

        for i, chart_row in enumerate(data["monthly_cancellation_chart"]):
            for col_idx, key in enumerate(["month", "cancelled", "refunded"], 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = chart_row.get(key, "")
                cell.border = border
                
                if col_idx == 1:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = right_aligned
                    
                if i % 2 == 1:
                    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.fill = fill
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="cancellation_report.xlsx"'
        return response


@extend_schema(tags=["Analytics"], summary="Export cancellation report", parameters=_PARAMS)
class CancellationReportExportView(APIView, _CancellationExportMixin):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_cancellation_report(date_from, date_to, currency=currency)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)


@extend_schema(tags=["Warehouse Analytics"], summary="Export warehouse cancellation report", parameters=_PARAMS)
class WarehouseCancellationReportExportView(APIView, _CancellationExportMixin):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        currency = request.query_params.get("currency", "").upper() or None
        if currency and currency not in CurrencyChoices.values:
            return Response({"detail": "Invalid currency."}, status=400)

        export_format = request.query_params.get("format", "csv").lower()
        if export_format not in ["csv", "excel"]:
            export_format = "csv"

        date_from, date_to = _parse_date_range(request)
        data = _build_cancellation_report(date_from, date_to, currency=currency, warehouse=request.warehouse)

        if export_format == "excel":
            return self._export_excel(data)
        else:
            return self._export_csv(data)
